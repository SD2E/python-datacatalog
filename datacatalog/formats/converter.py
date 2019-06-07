import json
import os
import sys
import inspect
from openpyxl import load_workbook
from shutil import copyfile
from jsonschema import validate, FormatChecker, ValidationError
from ..tenancy import Projects
# from .runner import convert_file

class ConversionError(Exception):
    """Something happened that prevented conversion of the target document"""
    pass

class formatChecker(FormatChecker):
    """A simple JSON format validator"""

    def __init__(self):
        FormatChecker.__init__(self)

class Converter(object):
    """Base class implementing a document converter"""
    VERSION = '0.0.0'
    FILENAME = 'baseclass'
    # Implementing subclasses should override
    projects = Projects.sync()
    PROJECT = projects.SD2.tacc_name
    TENANT = projects.SD2.tenant
    def __init__(self, schemas=[], targetschema=None, options={}, reactor=None):

        # Discover the default input schema
        HERE = os.path.abspath(inspect.getfile(self.__class__))
        PARENT = os.path.dirname(HERE)
        schema_path = os.path.join(PARENT, 'schema.json')
        # Input schema(s)
        # FIXME move to a single schema definition per class
        self.schemas = [schema_path]
        self.name = type(self).__name__
        if isinstance(schemas, str):
            if os.path.exists(str):
                self.schemas.append(str)
            else:
                raise OSError('schema file {} not found'.format(str))
        else:
            for s in schemas:
                if os.path.exists(s):
                    self.schemas.append(s)
                else:
                    raise OSError('schema file {} not found'.format(str))

        # Default output schema
        if targetschema is None:
            self.targetschema = { "$ref" : "https://schema.catalog.sd2e.org/schemas/sample_set.json" }
        else:
            self.targetschema = targetschema

        self.options = options
        self.reactor = reactor

        # Schema metadata
        setattr(self, 'filename', self.FILENAME)
        setattr(self, 'version', self.VERSION)
        setattr(self, 'project', self.PROJECT)
        setattr(self, 'tenant', self.TENANT)

    def convert(self, input_fp, output_fp=None, verbose=True, config={}, enforce_validation=True):
        """Convert between formats

        This is a pass-through method that invokes a runner script

        Args:
            input_fp (str): Path to input file
            output_fp (str): Path to output file
            verbose (bool, optional): Print verbose output while running
            config (dict, optional): Generic configuration object
            enforce_validation (bool, optional): Whether to force validation of outputs

        Returns:
            bool: Whether the conversion succeeeded
        """
        # Import lazily because of the SBH requirement
        from .runner import convert_file
        return convert_file(self.targetschema, input_fp, output_path=output_fp, verbose=verbose, config=config, enforce_validation=enforce_validation)

    def test(self, input_fp, output_fp, verbose=True, config={}):
        """Smoketest method to see if Converter discovery is working

        Returns:
            True
        """
        return True

    def validate_input(self, input_fp, encoding, permissive=False):
        """Validate a generic input file against schemas known to Converter

        Parameters:
            input_fp (str): path to the validation target file

        Arguments:
            permissive (bool): whether to return False on failure to validate

        Raises:
            ConversionError: Raised when schema or target can't be loaded
            ValidationError: Raised when validation fails

        Returns:
            boolean: True on success
        """
        # set encoding
        self.encoding = encoding

        # JSON Path
        if input_fp.endswith(".json"):
            try:
                with open(input_fp, 'r', encoding=encoding) as jsonfile:
                    jsondata = json.load(jsonfile)
            except Exception as exc:
                raise ConversionError('Failed to load {} for validation'.format(input_fp), exc)

            # Iterate through our schemas
            validation_errors = []
            for schema_path in self.schemas:
                try:
                    with open(schema_path) as schema:
                        schema_json = json.loads(schema.read())
                except Exception as e:
                    raise ConversionError(
                        'Failed to load schema for validation', e)

                try:
                    validate(jsondata, schema_json, format_checker=formatChecker())
                    return True
                except ValidationError as v:
                    validation_errors.append(v)
                    pass
                except Exception as e:
                    raise ConversionError(e)

            # If we have not returned True, all schemas failed
            if permissive:
                return False
            else:
                raise ValidationError(validation_errors)

        #XLSX PATH
        elif input_fp.endswith(".xlsx"):
            try:
                # load all sheets
                wb = load_workbook(input_fp, read_only=True)
            except Exception as exc:
                raise ConversionError('Failed to load {} for validation'.format(input_fp), exc)

            # Iterate through our schemas
            validation_errors = []
            for schema_path in self.schemas:
                try:
                    with open(schema_path) as schema:
                        schema_json = json.loads(schema.read())
                except Exception as e:
                    wb.close()
                    raise ConversionError(
                        'Failed to load schema for validation', e)

                try:
                    # pull headers from schema and check
                    schema_properties = schema_json["properties"]
                    if "xlsx" in schema_properties and schema_properties["xlsx"] and "headers" in schema_properties:

                        header_values_list = schema_properties["headers"]["oneOf"]

                        for header_values in header_values_list:
                            enum_values = [enum_item["enum"][0] for enum_item in header_values["items"]]
                            for sheetname in wb.sheetnames:
                                ws = wb[sheetname]
                                rows = ws.iter_rows(min_row=1, max_row=1)
                                first_row = next(rows)
                                excel_headers = [c.value for c in first_row]
                                valid = all([header in excel_headers for header in enum_values])
                                if valid:
                                    return valid
                except Exception as e:
                    wb.close()
                    raise ConversionError(e)

            wb.close()

            # If we have not returned True, all schemas failed
            if permissive:
                return False
            else:
                raise ValidationError(validation_errors)

    def validate(self, output_fp, permissive=False):
        """Validate a file against schemas known to Converter

        Parameters:
            output_fp (str): path to the validation target file

        Note:
            Yes, this is redundant with validate_input()

        Arguments:
            permissive (bool): whether to return False on failure to validate

        Raises:
            ValidationError: Raised when validation fails

        Returns:
            boolean: True on success
        """
        try:
            with open(output_fp, 'r') as jsonfile:
                jsondata = json.load(jsonfile)
        except Exception as exc:
            raise ValidationError(
                'Unable to load {} for validation'.format(output_fp), exc)

        try:
            with open(self.targetschema) as schema:
                schema_json = json.loads(schema.read())
        except Exception as e:
            raise ValidationError('Unable to load schema for validation', e)

        try:
            validate(jsondata, schema_json, format_checker=formatChecker())
            return True
        except ValidationError as v:
            if permissive:
                return False
            else:
                raise ValidationError('Schema validation failed', v)
        except Exception as e:
            raise ValidationError(e)

    def get_schema(self):
        """Pass-through to ``get_classifier_schema()``"""
        return self.get_classifier_schema()

    def get_classifier_schema(self):
        """Get the JSON schema that Converter is using for classification

        Raises:
            ConversionError: Returned on all Exceptions

        Returns:
            dict: JSON schema in dictionary form
        """
        # Return the classifier schema as a Python object
        schema_fp = getattr(self, 'schemas', [])[0]
        try:
            with open(schema_fp, 'r') as jsonfile:
                return json.load(jsonfile)
        except Exception as exc:
            raise ConversionError('Failed to load {}'.format(schema_fp), exc)
